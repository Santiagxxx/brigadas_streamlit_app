from __future__ import annotations

from copy import copy
from io import BytesIO
from pathlib import Path
from typing import Dict, Iterable, List

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

BASE_DIR = Path(__file__).resolve().parents[1]
TEMPLATE_PATH = BASE_DIR / "templates" / "FORMATOBRIGADAS.xlsx"

HEADERS = [
    "CODIGO MCA",
    "NOMBRE MCA",
    "BRIGADAS REALIZADAS",
    "# PERSONAS BRIGADA 1",
    "# PERSONAS BRIGADA 2",
    "REGION",
]

EXPORT_FILE_NAME = "FORMATO_BRIGADAS_RESPUESTAS.xlsx"


def _capture_body_style(ws: Worksheet, template_row: int = 2) -> List[dict]:
    styles: List[dict] = []
    for col in range(1, len(HEADERS) + 1):
        source = ws.cell(template_row, col)
        styles.append(
            {
                "font": copy(source.font),
                "fill": copy(source.fill),
                "border": copy(source.border),
                "alignment": copy(source.alignment),
                "number_format": source.number_format,
                "protection": copy(source.protection),
            }
        )
    return styles


def _apply_style(cell, style: dict) -> None:
    cell.font = copy(style["font"])
    cell.fill = copy(style["fill"])
    cell.border = copy(style["border"])
    cell.alignment = copy(style["alignment"])
    cell.number_format = style["number_format"]
    cell.protection = copy(style["protection"])


def _prepare_sheet(ws: Worksheet, body_styles: List[dict], body_height) -> None:
    # Mantiene la hoja y encabezados del formato original.
    for col, header in enumerate(HEADERS, start=1):
        ws.cell(1, col).value = header

    # Limpia únicamente el cuerpo A:F, sin tocar estilos de encabezado.
    max_row = max(ws.max_row, 2)
    for row in range(2, max_row + 1):
        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(row, col)
            cell.value = None
            _apply_style(cell, body_styles[col - 1])
        if body_height:
            ws.row_dimensions[row].height = body_height

    ws.freeze_panes = "A2"


def build_excel_bytes(records: Iterable[Dict[str, object]], template_path: Path = TEMPLATE_PATH) -> bytes:
    """Builds an XLSX with the original FORMATOBRIGADAS.xlsx layout."""
    if not template_path.exists():
        raise FileNotFoundError(f"No se encontró la plantilla: {template_path}")

    wb = load_workbook(template_path)
    ws = wb.active

    # Capturar estilos antes de limpiar la fila de datos.
    body_styles = _capture_body_style(ws, template_row=2)
    body_height = ws.row_dimensions[2].height

    _prepare_sheet(ws, body_styles, body_height)

    last_row = 1
    for row_idx, record in enumerate(records, start=2):
        values = [
            record.get("codigo_mca", ""),
            record.get("nombre_mca", ""),
            int(record.get("brigadas_realizadas", 0) or 0),
            int(record.get("personas_brigada_1", 0) or 0),
            int(record.get("personas_brigada_2", 0) or 0),
            record.get("region", ""),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row_idx, col_idx)
            cell.value = value
            _apply_style(cell, body_styles[col_idx - 1])
        if body_height:
            ws.row_dimensions[row_idx].height = body_height
        last_row = row_idx

    # Mantiene filtros simples para facilitar revisión desde Excel.
    ws.auto_filter.ref = f"A1:F{max(last_row, 2)}"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def save_excel_file(records: Iterable[Dict[str, object]], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_bytes(build_excel_bytes(records))
    return output_path
